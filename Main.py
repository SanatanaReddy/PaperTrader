import openpyxl
import sys
import matplotlib.lines as lin
import matplotlib.text as text
from matplotlib.patches import Rectangle
import matplotlib.pyplot as plt
import pandas as pd
import yfinance as yf
	
def nextCandlestick(positionAfterOpen = 0,width=0.6, colorup='g', colordown='r',alpha=1.0):
	global ax,ws,wb,lineMode,figureTitle,titleText,candlenumber, currentPosition, entryPrice,entryDate, PandL, datesList,opensList, closesList, highsList, lowsList
	OFFSET = width/2.0
	lineMode = False
	close = closesList[candlenumber]
	open = opensList[candlenumber]
	if currentPosition != 0 and currentPosition != positionAfterOpen:
		profit = currentPosition * (open - entryPrice)
		PandL += profit
		ws.append((positionStrings[currentPosition],entryPrice,open,(datesList[candlenumber]-entryDate).days,profit))
		wb.save('/Users/sanatana/Documents/Paper_Trades.xlsx')
	if positionAfterOpen != 0 and currentPosition != positionAfterOpen:
		entryPrice = open
		entryDate = datesList[candlenumber]
	currentPosition = positionAfterOpen
	tradeValue = currentPosition * (close - entryPrice)
	if close >= open:
		color = colorup
		lower = open
		height = close - open
	else:
		color = colordown
		lower = close
		height = open - close
	vline = lin.Line2D(xdata=(candlenumber+1, candlenumber+1), ydata=(lowsList[candlenumber], highsList[candlenumber]),color=color, linewidth=1.0, antialiased=True)
	rect = Rectangle(xy=(candlenumber +1 - OFFSET, lower),width=width,height=height,facecolor=color,edgecolor=color )
	rect.set_alpha(alpha)
	ax.add_line(vline)
	ax.add_patch(rect)
	setView()
	candlenumber += 1
	if currentPosition == 0:
		entryString = "---"
		valueString = "---"
	else:
		entryString = str(round(entryPrice,2))
		valueString = str(round(tradeValue,2))
	titleText = "P&L: " + str(round(PandL,2)) +  "      Currently: " + positionStrings[currentPosition]
	
	titleText +=  "     Entry: " + entryString + "     Latest Price: " + str(round(close,2)) + "    Position Value: " + valueString
	figureTitle.set_text(titleText)


def press(event):
	global candlestickCounter,candlenumber, lineMode,drawnLines,figureTitle,titleText
	sys.stdout.flush()
	if event.key == 'left':
		nextCandlestick()
		fig.canvas.draw()
	if event.key == 'right':
		nextCandlestick(currentPosition)
		fig.canvas.draw()
	if event.key == 'up':
		nextCandlestick(positionAfterOpen = 1)	
		fig.canvas.draw()
	if event.key == 'down':
		nextCandlestick(positionAfterOpen = -1)
		fig.canvas.draw()
	if event.key == 'd':
		lineMode = not lineMode
		if lineMode:
			figureTitle.set_text("CLICK ON TWO POINTS TO DRAW A LINE")
		else:
			figureTitle.set_text(titleText)
		fig.canvas.draw()
	if event.key == 'w':
		if len(drawnLines) == 0:
			return
		ax.lines.remove(drawnLines[len(drawnLines)-1])
		drawnLines.pop()
		fig.canvas.draw()
	if event.key == 'q':
		summarizeTrades()

def click(event):
	global firstXCoor, firstYCoor, drawingLine, lineMode
	sys.stdout.flush()
	if (not lineMode) or event.inaxes != ax:
		return
	if not drawingLine:
		drawingLine = True
		firstXCoor = event.xdata
		firstYCoor = event.ydata
	else:
		drawLine(firstXCoor,firstYCoor, event.xdata,event.ydata)
		drawingLine = False

def drawLine(x1,y1,x2,y2):
	global drawnLines
	if x1 == x2:
		extendedXData = [x1,x2]
		extendedYData = [0,10*y2]
	else:
		slope = (y2-y1)/(x2-x1)
		intercept = y2 - (slope*x2)
		extendedXData = [0,numberOfDays]
		extendedYData = [intercept, slope*numberOfDays + intercept]
	drawnLines.append(ax.plot(extendedXData,extendedYData,scalex = False,scaley = True,color = 'b')[0])
	fig.canvas.draw()

def setView():
	global candlenumber,highsList,lowsList,lowerYView,upperYView
	if candlenumber >= 200:
		begin = candlenumber - 200
		end  = candlenumber
	else: 
		begin = 0
		end = candlenumber
	ax.set_xlim(left = begin, right = end + 15)
	if candlenumber%5 != 0 and highsList[candlenumber] < 0.95 * upperYView and lowsList[candlenumber] > 1.05* lowerYView:
		ax.set_ylim(bottom = lowerYView, top = upperYView)
		return
	lowest = lowsList[begin]
	highest = highsList[end]
	for i in range(candlenumber - 200, candlenumber+1):
		lowest = min(lowest,lowsList[i])
		highest = max(highest, highsList[i])
	lowerYView = lowest - (highest-lowest)*0.2
	upperYView = highest + (highest - lowest)*0.2
	ax.set_ylim(bottom = lowerYView, top = upperYView)

def summarizeTrades():
	global ws, wb
	if ws.max_row > 1:
		for row in range(2,ws.max_row + 1):
			ws.cell(row,6).value = ws.cell(row,5).value/ws.cell(row,2).value*100
			ws.cell(row,7).value = ws.cell(row-1,7).value + ws.cell(row,5).value
	ws.cell(1,7).value = 'P & L'
	wb.save('/Users/sanatana/Documents/Paper_Trades.xlsx')

currentPosition = 0
PandL = 0
entryPrice = 0
entryDate = 0
candlenumber = 0
tradeValue = 0
firstXCoor = 0
firstYCoor = 0
lowerYView = 0
upperYView = 0
titleText = ""
drawingLine = False
lineMode = False
drawnLines = []
positionStrings = {
	-1: "Short",
	0: "Neutral",
	1: "Long"
}


wb = openpyxl.load_workbook('/Users/sanatana/Documents/Paper_Trades.xlsx')
symbol = input("Type a symbol in uppercase letters:\n")
ws = wb.create_sheet(title = symbol)
ws.append(('Direction','Entry Price($)','Exit Price($)','Trade Duration(days)','Profit($)','Profit(%)',0))
data = yf.Ticker(symbol)
quotes = data.history(start = "2013-07-07",end = "2018-07-07")
datesList = quotes.index.tolist()
opensList = quotes['Open'].tolist()
highsList = quotes['High'].tolist()
lowsList = quotes['Low'].tolist()
closesList = quotes['Close'].tolist()
numberOfDays = len(datesList)

fig, ax = plt.subplots()
fig.subplots_adjust(left = 0.05,bottom=0.05, right = 0.95, top = 0.95)
ax.set_autoscalex_on(False)
ax.set_autoscaley_on(False)
lowerYView = lowsList[0]
ax.set_xlabel("Days")
ax.set_facecolor((0.96,0.95,0.92))
figureTitle = plt.title("",loc = 'left')
for i in range(0,201):
	nextCandlestick()


plt.setp(plt.gca().get_xticklabels(), rotation=45, horizontalalignment='right')
fig.canvas.set_window_title(symbol)
fig.canvas.mpl_connect('key_press_event', press)
fig.canvas.mpl_connect('button_press_event', click)
plt.grid(linestyle = '--')
plt.show()
